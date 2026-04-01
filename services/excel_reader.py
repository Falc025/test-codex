from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from models.document_data import DocumentData
from services.validator import InputValidator, ValidationError


class ExcelReader:
    """Lee registros desde Excel tabular y produce una lista de DocumentData."""

    def __init__(self, validator: InputValidator | None = None) -> None:
        self.validator = validator or InputValidator()

    def read_document_data(self, excel_path: str | Path) -> list[DocumentData]:
        validated_path = self.validator.validate_excel_path(excel_path)

        try:
            workbook = load_workbook(validated_path, data_only=True)
        except Exception as exc:
            raise ValidationError(f"No se pudo abrir el Excel: {exc}") from exc

        self.validator.validate_sheet_exists(workbook.sheetnames)
        sheet = workbook[InputValidator.REQUIRED_SHEET]

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("La hoja 'Datos' está vacía.")

        headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        self.validator.validate_header_columns(headers)

        records: list[DocumentData] = []
        for idx, row in enumerate(rows[1:], start=2):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            raw = {headers[col_idx]: value for col_idx, value in enumerate(row) if col_idx < len(headers) and headers[col_idx]}
            self.validator.validate_row_data(raw, idx)
            records.append(DocumentData.from_raw(raw, row_number=idx))

        if not records:
            raise ValidationError("No se encontraron registros válidos en la hoja 'Datos'.")

        return records
