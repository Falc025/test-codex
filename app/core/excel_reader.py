from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format

from app.utils.text_utils import normalize_column_name


@dataclass
class ExcelRow:
    raw: dict[str, Any]
    display: dict[str, str]
    source_row: int


@dataclass
class ExcelData:
    rows: list[ExcelRow]
    columns: list[str]
    sheet_names: list[str]
    selected_sheet: str
    warnings: list[str]


class ExcelReader:
    MONTH_ABBR_ES = {
        1: "Ene",
        2: "Feb",
        3: "Mar",
        4: "Abr",
        5: "May",
        6: "Jun",
        7: "Jul",
        8: "Ago",
        9: "Sep",
        10: "Oct",
        11: "Nov",
        12: "Dic",
    }
    MONTH_NAME_ES = {
        1: "Enero",
        2: "Febrero",
        3: "Marzo",
        4: "Abril",
        5: "Mayo",
        6: "Junio",
        7: "Julio",
        8: "Agosto",
        9: "Septiembre",
        10: "Octubre",
        11: "Noviembre",
        12: "Diciembre",
    }

    def get_sheet_names(self, excel_path: str | Path) -> list[str]:
        wb = load_workbook(excel_path, data_only=True)
        return wb.sheetnames

    def read(self, excel_path: str | Path, sheet_name: str | None = None) -> ExcelData:
        wb_data = load_workbook(excel_path, data_only=True)
        wb_formula = load_workbook(excel_path, data_only=False)

        selected = sheet_name or wb_data.sheetnames[0]
        ws_data = wb_data[selected]
        ws_formula = wb_formula[selected]

        header_row = next(ws_data.iter_rows(min_row=1, max_row=1))
        headers: list[str] = [normalize_column_name(cell.value) for cell in header_row]

        rows: list[ExcelRow] = []
        warnings: list[str] = []
        for row_idx, row_cells in enumerate(ws_data.iter_rows(min_row=2), start=2):
            row_raw: dict[str, Any] = {}
            row_display: dict[str, str] = {}
            has_content = False

            for i, cell in enumerate(row_cells):
                if i >= len(headers):
                    continue
                col = headers[i]
                if not col:
                    continue
                raw_val = cell.value
                disp_val = self.excel_display_value(cell)
                row_raw[col] = raw_val
                row_display[col] = disp_val
                if disp_val != "" or raw_val is not None:
                    has_content = True

                formula_cell = ws_formula.cell(row=row_idx, column=i + 1)
                if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") and raw_val is None:
                    warnings.append(
                        f"Hoja {selected} fila {row_idx} columna '{col}': fórmula sin valor calculado guardado"
                    )

            if has_content:
                rows.append(ExcelRow(raw=row_raw, display=row_display, source_row=row_idx))

        return ExcelData(
            rows=rows,
            columns=[h for h in headers if h],
            sheet_names=wb_data.sheetnames,
            selected_sheet=selected,
            warnings=warnings,
        )

    def excel_display_value(self, cell) -> str:
        raw = cell.value
        if raw is None:
            return ""

        if isinstance(raw, str):
            return raw.strip()

        if isinstance(raw, bool):
            return "TRUE" if raw else "FALSE"

        fmt = (cell.number_format or "").strip()

        if isinstance(raw, (datetime, date)) or (is_date_format(fmt) and isinstance(raw, (int, float))):
            dt = raw if isinstance(raw, (datetime, date)) else self._excel_serial_to_datetime(raw)
            return self.format_excel_date(dt, fmt)

        if isinstance(raw, (int, float, Decimal)):
            return self._format_number(float(raw), fmt)

        return str(raw).strip()

    def _format_number(self, value: float, number_format: str) -> str:
        fmt = (number_format or "General").split(";")[0]

        if "%" in fmt:
            decimals = self._count_decimals(fmt)
            scaled = value * 100
            # Mantener formato porcentual existente, con conversión decimal a coma institucional.
            return self.format_number_es(scaled, decimals=decimals, use_thousands=False) + "%"

        decimals = self._count_decimals(fmt)
        if decimals == 0:
            decimals = 2
        use_thousands = "#" in fmt and any(sep in fmt for sep in ("#,##", "# ##", "#.##"))
        return self.format_number_es(value=value, decimals=decimals, use_thousands=use_thousands)

    @staticmethod
    def _count_decimals(fmt: str) -> int:
        clean_fmt = (fmt or "").split(";", 1)[0]
        matches = re.findall(r"[.,]([0#]+)", clean_fmt)
        if not matches:
            return 0
        return len(matches[-1])

    @staticmethod
    def format_number_es(value: float | int | Decimal, decimals: int = 2, use_thousands: bool = True) -> str:
        if value is None:
            return ""
        try:
            dec_value = Decimal(str(value))
        except Exception:
            return str(value).strip()

        quant = Decimal(f"1.{'0' * decimals}") if decimals > 0 else Decimal("1")
        quantized = dec_value.quantize(quant, rounding=ROUND_HALF_UP)
        sign = "-" if quantized < 0 else ""
        normalized = format(abs(quantized), f".{decimals}f")
        integer_part, decimal_part = normalized.split(".") if "." in normalized else (normalized, "")

        if use_thousands:
            groups: list[str] = []
            for i in range(len(integer_part), 0, -3):
                groups.append(integer_part[max(0, i - 3):i])
            integer_part = " ".join(reversed(groups))

        if decimals > 0:
            return f"{sign}{integer_part},{decimal_part}"
        return f"{sign}{integer_part}"

    @staticmethod
    def _excel_serial_to_datetime(value: float) -> datetime:
        base = datetime(1899, 12, 30)
        return base + timedelta(days=float(value))

    def format_excel_date(self, value: date | datetime, number_format: str) -> str:
        dt = value if isinstance(value, datetime) else datetime.combine(value, datetime.min.time())
        fmt = self._clean_excel_date_format(number_format)
        lower_fmt = fmt.lower()

        if re.fullmatch(r"d{1,2}/m{1,2}/y{4}", lower_fmt):
            day = str(dt.day) if lower_fmt.startswith("d/") else f"{dt.day:02d}"
            month = str(dt.month) if "/m/" in lower_fmt else f"{dt.month:02d}"
            return f"{day}/{month}/{dt.year:04d}"
        if re.fullmatch(r"mm/yyyy", lower_fmt):
            return f"{dt.month:02d}/{dt.year:04d}"
        if re.fullmatch(r"mmm-yy", lower_fmt):
            return f"{self._month_abbr(dt.month)}-{dt.year % 100:02d}"
        if re.fullmatch(r"mmmm-yy", lower_fmt):
            return f"{self._month_name(dt.month)}-{dt.year % 100:02d}"
        if re.fullmatch(r"mmm/yyyy", lower_fmt):
            return f"{self._month_abbr(dt.month)}/{dt.year:04d}"
        if re.fullmatch(r"mm-yy", lower_fmt):
            return f"{dt.month:02d}-{dt.year % 100:02d}"
        if re.fullmatch(r"dd/mm/yyyy", lower_fmt):
            return f"{dt.day:02d}/{dt.month:02d}/{dt.year:04d}"
        if re.fullmatch(r"d/m/yyyy", lower_fmt):
            return f"{dt.day}/{dt.month}/{dt.year:04d}"

        return dt.strftime("%d/%m/%Y")

    @staticmethod
    def _clean_excel_date_format(number_format: str) -> str:
        base = (number_format or "").split(";")[0].strip()
        base = re.sub(r"\[[^\]]*\]", "", base)
        base = base.replace("\\", "")
        base = re.sub(r'"[^"]*"', "", base)
        return base.strip()

    @staticmethod
    def _month_abbr(month: int) -> str:
        return ExcelReader.MONTH_ABBR_ES[month]

    @staticmethod
    def _month_name(month: int) -> str:
        return ExcelReader.MONTH_NAME_ES[month]
